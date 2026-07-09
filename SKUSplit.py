#!/usr/bin/env python3
"""
SKU店铺拆分工具 — 模拟退火 + 定向交换
先用多轮随机选最佳初始解，再用模拟退火逐步降低最大重叠。
对重复率使用严格小于 (<n%)。
"""

import sys, os, math, random, time, warnings
from collections import Counter
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# 抑制openpyxl在frozen exe中的无关警告
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

TIME_LIMIT = 300

# 获取程序自身所在目录（兼容frozen exe）
def get_script_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(os.path.abspath(sys.executable))
    else:
        return os.path.dirname(os.path.abspath(__file__))

# ============ 基础函数 ============

def read_input(fp):
    wb = openpyxl.load_workbook(fp)
    rp = list(wb['参数表'].iter_rows(values_only=True))
    dr = [r for r in rp if r[0] is not None and not str(r[0]).strip().startswith('店铺')]
    ns = int(dr[0][0]); sps = int(dr[0][1]); mor = float(dr[0][2])
    mo = math.ceil(sps * mor) - 1  # 严格小于 <n%
    sd = []
    for row in wb['sku表'].iter_rows(min_row=2, values_only=True):
        sk = str(row[0]).strip() if row[0] is not None else ''
        if not sk or sk == 'None' or sk == '': continue
        price = row[2] if len(row) > 2 else None
        weight = row[4] if len(row) > 4 else None
        sd.append((sk, price, weight))
    wb.close()
    return ns, sps, mor, mo, sd


def gen(ns, sps, nu, seed):
    random.seed(seed)
    return [set(random.sample(range(nu), sps)) for _ in range(ns)]


def max_overlap(stores, ns):
    mo = 0
    for i in range(ns):
        si = stores[i]
        for j in range(i+1, ns):
            ov = len(si & stores[j])
            if ov > mo: mo = ov
    return mo


def compute_overlap_matrix(stores, ns):
    om = [[0]*ns for _ in range(ns)]
    for i in range(ns):
        si = stores[i]
        for j in range(i+1, ns):
            o = len(si & stores[j])
            om[i][j] = o; om[j][i] = o
    return om


def find_best_start(ns, sps, nu, t0, n_trials=2000):
    best = float('inf')
    best_stores = None
    best_seed = None
    for trial in range(1, n_trials + 1):
        if time.time() - t0 > TIME_LIMIT * 0.3: break
        stores = gen(ns, sps, nu, trial * 137)
        m = max_overlap(stores, ns)
        if m < best:
            best = m
            best_stores = [set(s) for s in stores]
            best_seed = trial
            if trial % 200 == 0:
                print(f"  随机试#{trial}: 当前最佳={best}")
    print(f"  最佳初始解: 最大重叠={best} (seed={best_seed})")
    return best_stores, best


def simulated_annealing(stores, ns, sps, mo_target, nu, t0):
    """
    模拟退火优化。目标: 每对店重叠 ≤ mo_target。
    每次选最严重的超限店对(i,j)，在i中替换一个共同SKU。
    """
    om = compute_overlap_matrix(stores, ns)
    
    cur_max = max(om[i][j] for i in range(ns) for j in range(i+1, ns))
    best_max = cur_max
    best_stores = [set(s) for s in stores]
    
    T = 8.0; T_min = 0.01; cooling = 0.999
    stall = 0
    last_improvement = time.time()
    restart_count = 0
    
    while T > T_min and (time.time() - t0 < TIME_LIMIT * 0.95):
        viol_pairs = [(i, j, om[i][j]) for i in range(ns) for j in range(i+1, ns) if om[i][j] > mo_target]
        
        if not viol_pairs:
            return True, stores
        
        # 95%概率选最严重对，5%随机选
        i, j, ov = max(viol_pairs, key=lambda x: x[2]) if random.random() < 0.95 else random.choice(viol_pairs)
        
        common = list(stores[i] & stores[j])
        if not common: continue
        sku = random.choice(common)
        
        candidates = [s for s in range(nu) if s not in stores[i]]
        random.shuffle(candidates)
        
        found = False
        for new_sk in candidates[:50]:
            # 只检查受影响的店
            delta_max = 0
            for k in range(ns):
                if k == i: continue
                delta = 0
                if sku in stores[k]: delta -= 1
                if new_sk in stores[k]: delta += 1
                nv = om[i][k] + delta
                if nv > delta_max: delta_max = nv
            
            old_max_i = max(om[i][k] for k in range(ns) if k != i)
            delta_E = delta_max - old_max_i
            
            if delta_E < 0 or random.random() < math.exp(-delta_E / max(T, 0.001)):
                stores[i].discard(sku); stores[i].add(new_sk)
                for k in range(ns):
                    if k == i: continue
                    d = 0
                    if sku in stores[k]: d -= 1
                    if new_sk in stores[k]: d += 1
                    om[i][k] += d; om[k][i] += d
                found = True
                break
        
        if found:
            new_max = max(om[a][b] for a in range(ns) for b in range(a+1, ns))
            if new_max < best_max:
                best_max = new_max
                best_stores = [set(s) for s in stores]
                stall = 0; last_improvement = time.time()
        else:
            stall += 1
        
        T *= cooling
        
        if stall > 10000 and time.time() - last_improvement > 20:
            stores = [set(s) for s in best_stores]
            om = compute_overlap_matrix(stores, ns)
            T = 4.0
            restart_count += 1
            stall = 0
            print(f"    退火重启#{restart_count} (最佳={best_max})")
    
    return False, best_stores


# ============ 输出 ============

def write_output(fp, stores, uniq, data, ns, sps, mor, om):
    wb = openpyxl.Workbook()
    mo = math.ceil(sps * mor) - 1
    ws = wb.active; ws.title = '汇总'
    all_o = [om[i][j] for i in range(ns) for j in range(i+1, ns)]
    mx = max(all_o) if all_o else 0; avg = sum(all_o)/len(all_o) if all_o else 0
    fc = Counter()
    for s in stores:
        for sk in s: fc[sk] += 1
    fd = sorted(Counter(fc.values()).items())
    
    rows = [
        ('====== 参数 ======',''), ('店铺数量',ns), ('每店SKU数',sps),
        ('重复率上限(严格小于)',f'{mor:.0%}'), ('允许最大重叠SKU',mo),
        ('',''), ('====== 结果统计 ======',''),
        ('唯一SKU总数',len(uniq)), ('总Slot数',ns*sps),
        ('实际最大两两重叠',f'{mx} ({mx/sps:.1%})'),
        ('实际平均两两重叠',f'{avg:.1f} ({avg/sps:.1%})'),
        ('',''), ('====== SKU频率分布 ======',''),
    ]
    for f,c in fd: rows.append((f'  出现{f}次',f'{c}个'))
    
    hf = PatternFill(start_color='4472C4',end_color='4472C4',fill_type='solid')
    hn = Font(bold=True,color='FFFFFF',size=11)
    for r,(l,v) in enumerate(rows,1):
        c1=ws.cell(row=r,column=1,value=l); ws.cell(row=r,column=2,value=v)
        if '==' in str(l): c1.font=Font(bold=True,size=13)
        elif l and not l.startswith('  '): c1.font=Font(bold=True)
    ws.column_dimensions['A'].width=28; ws.column_dimensions['B'].width=22
    
    for i in range(ns):
        w=wb.create_sheet(title=f'店铺{i+1}'[:31])
        for c,h in enumerate(['SKU','售价','重量'],1):
            cell=w.cell(row=1,column=c,value=h); cell.fill=hf; cell.font=hn; cell.alignment=Alignment(horizontal='center')
        for r,si in enumerate(sorted(stores[i]),2):
            sk,pr,wt=data[si]; w.cell(row=r,column=1,value=sk); w.cell(row=r,column=2,value=pr); w.cell(row=r,column=3,value=wt)
        w.column_dimensions['A'].width=16; w.column_dimensions['B'].width=10; w.column_dimensions['C'].width=10
    
    ws2=wb.create_sheet(title='店铺间重复率')
    red=PatternFill(start_color='FF6B6B',end_color='FF6B6B',fill_type='solid')
    grn=PatternFill(start_color='90EE90',end_color='90EE90',fill_type='solid')
    yel=PatternFill(start_color='FFD700',end_color='FFD700',fill_type='solid')
    ws2.cell(row=1,column=1).fill=hf
    for j in range(ns):
        c=ws2.cell(row=1,column=j+2,value=f'店{j+1}'); c.fill=hf; c.font=hn; c.alignment=Alignment(horizontal='center')
    for i in range(ns):
        ws2.cell(row=i+2,column=1,value=f'店{i+1}').font=Font(bold=True)
        for j in range(ns):
            if i==j: ws2.cell(row=i+2,column=j+2,value='-').alignment=Alignment(horizontal='center')
            else:
                ov=om[i][j]; rt=ov/sps
                c=ws2.cell(row=i+2,column=j+2,value=round(rt,4))
                c.number_format='0.00%'; c.alignment=Alignment(horizontal='center')
                c.fill=red if rt>mor else (yel if rt>mor*0.9 else grn)
    ws2.freeze_panes='B2'; ws2.column_dimensions['A'].width=10
    for col in range(2,ns+2): ws2.column_dimensions[get_column_letter(col)].width=10
    wb.save(fp)


# ============ 主流程 ============

def main():
    sd = get_script_dir()
    inp=os.path.join(sd,'sku数据.xlsx')
    if not os.path.exists(inp): print("错误: 找不到 sku数据.xlsx"); sys.exit(1)
    out=os.path.join(sd,'店铺拆分结果.xlsx')
    print(f"读取: {inp}")
    ns,sps,mor,mo,sd_data=read_input(inp)
    seen={}; deduped=[]
    for sk,pr,wt in sd_data:
        if sk not in seen: seen[sk]=len(deduped); deduped.append((sk,pr,wt))
    nu=len(deduped); total=ns*sps
    print(f"\n店铺: {ns}, 每店SKU: {sps}, 重复率 < {mor:.0%} (严格)")
    print(f"即: 两两重叠必须 ≤ {mo} 个SKU")
    print(f"唯一SKU: {nu}, 总Slot: {total}, 平均: {total/nu:.2f}次/个")
    t0 = time.time()
    
    # 可行性检查
    n_pairs = ns*(ns-1)//2; budget = n_pairs * mo
    base = total // nu; rem = total % nu
    min_ss = rem*(base+1)**2 + (nu-rem)*base**2; min_ov = (min_ss - total)/2
    exp_ov = nu * (total/nu) * ((total/nu)-1) / (ns*(ns-1))
    print(f"  期望重叠: {exp_ov:.1f}, 预算: {budget}, 最小理论总重叠: {min_ov:.0f}")
    if nu < sps: print(f"\n✗ 唯一SKU({nu}) < 每店SKU数({sps})"); sys.exit(1)
    if min_ov > budget:
        print(f"\n✗ 理论最小总重叠({min_ov:.0f}) > 预算({budget:.0f})，理论上无解"); sys.exit(1)
    
    # 阶段1: 最佳初始解
    print(f"\n=== 阶段1: 随机搜索最佳初始解 (≤{TIME_LIMIT*0.3:.0f}s) ===")
    stores, init_max = find_best_start(ns, sps, nu, t0)
    
    if init_max <= mo:
        print(f"初始解已满足约束!")
        om = compute_overlap_matrix(stores, ns)
    else:
        print(f"\n=== 阶段2: 模拟退火优化 (需要从{init_max}降到{mo}) ===")
        ok, best_stores = simulated_annealing(stores, ns, sps, mo, nu, t0)
        if ok:
            print("  成功!")
        else:
            print("  退火结束")
        stores = best_stores
        om = compute_overlap_matrix(stores, ns)
    
    max_o = max(om[i][j] for i in range(ns) for j in range(i+1, ns))
    print(f"\n最终最大两两重叠: {max_o}/{mo} ({max_o/sps:.1%})")
    print(f"所有SKU数={sps}: {'✓' if all(len(s)==sps for s in stores) else '✗'}")
    print(f"所有重叠≤{mo}: {'✓' if max_o<=mo else '✗'}")
    
    if max_o > mo:
        print(f"\n⚠ 无法找到严格满足<{mor:.0%}的方案（当前最佳{max_o}/{sps}={max_o/sps:.1%}）")
        print(f"  建议: 增加唯一SKU数 / 减少店铺数 / 降低每店SKU数 / 提高重复率上限")
    else:
        print(f"\n✓ 全部约束满足!")
    
    print(f"\n写入: {out}")
    write_output(out, stores, [s[0] for s in deduped], deduped, ns, sps, mor, om)
    print(f"完成! ({time.time()-t0:.1f}s)")


if __name__=='__main__':
    main()
